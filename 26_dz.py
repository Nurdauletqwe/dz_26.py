import tkinter as tk
from openpyxl import load_workbook
import os

class ExcelCounter:
    def __init__(self, root):
      self.root = root
      self.root.title("Подсчет строк в Excel")
      self.root.geometry("400x200")
      self.result_label = tk.Label(self.root, text="Общее количество строк:")
      self.result_label.pack()
      self.count_button = tk.Button(self.root, text="Подсчитать", command= self.update_result())
      self.count_button.pack()

def count_rows(self, path):
    total_rows = 0
    for filename in os.listdir(path):
        if filename.endswith(".xlsx"):
            workbook = load_workbook(os.path.join(path, filename))
            for sheet in workbook.worksheets:
                total_rows += sheet.max_row
    return total_rows

def update_result(self):
    path = "c:\Users\hawci\OneDrive\Рабочий стол\nurda\25_dz.py"
    total_rows = count_rows(path)
    self.result_label.config(text=f"Общее количество строк: {total_rows}")
    
def main():
    root = tk.Tk()
    app = ExcelCounter(root)
    root.mainloop()


if __name__ == "__main__":
    main()